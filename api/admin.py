import csv
import datetime
import io
import uuid
from collections import OrderedDict, defaultdict

from django.contrib import admin, messages
from django.contrib.auth.admin import GroupAdmin as BaseGroupAdmin
from django.contrib.auth.admin import UserAdmin
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.models import Group
from django.db.models import Count, Max, Sum
from django.http import StreamingHttpResponse, HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils import timezone
from django.utils.crypto import get_random_string
from django.utils.safestring import mark_safe
from import_export import resources
from import_export.admin import ImportExportModelAdmin
from import_export.formats import base_formats
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from twisted.protocols.wire import Echo
from unfold.admin import ModelAdmin, StackedInline
from unfold.contrib.filters.admin import RelatedDropdownFilter, ChoicesDropdownFilter
from unfold.decorators import action
from unfold.forms import AdminPasswordChangeForm, UserChangeForm, UserCreationForm

from .forms import ConstructionObjectForm, GenerateUsersForm
from .models import *
from .resources import LoginAttemptsResource


# ---------------------------------------------------------------------------
# ConstructionObject -> Neighborhood -> District svod hisobot (Excel)
# ---------------------------------------------------------------------------

def _build_objects_summary_by_district():
    """
    ConstructionObject larni District kesimida guruhlaydi va quyidagi
    ko'rsatkichlarni hisoblaydi:
      - jami obyektlar soni (Sum(ConstructionObject.building_count))
      - moliyalashtirish ko'rsatkichi past (< 15%) bo'lgan loyihalar soni
      - so'nggi kunlik ma'lumot 7 kundan oshib ketgan loyihalar soni
      - umuman kunlik ma'lumot kiritilmagan loyihalar soni
    """
    today = timezone.now().date()

    financing_totals = dict(
        ConstructionFinancing.objects.values('construction_id')
        .annotate(total=Sum('amount'))
        .values_list('construction_id', 'total')
    )
    last_progress_dates = dict(
        ConstructionDailyProgress.objects.values('construction_id')
        .annotate(last_date=Max('date'))
        .values_list('construction_id', 'last_date')
    )

    objects = ConstructionObject.objects.select_related('neighborhood__district')

    districts = OrderedDict()

    for obj in objects:
        neighborhood = obj.neighborhood
        district = neighborhood.district if neighborhood else None

        district_key = district.id if district else 0
        district_name = district.name if district else "Tuman biriktirilmagan"

        bucket = districts.setdefault(district_key, {
            'name': district_name,
            'total_buildings': 0,
            'low_financing': 0,
            'stale': 0,
            'no_data': 0,
        })

        bucket['total_buildings'] += obj.building_count or 0

        total_financed = financing_totals.get(obj.id) or 0
        budget = obj.budget or 0
        if budget > 0:
            ratio = (total_financed / budget) * 100
            if ratio < 15:
                bucket['low_financing'] += 1

        last_date = last_progress_dates.get(obj.id)
        if last_date is None:
            bucket['no_data'] += 1
        elif (today - last_date).days > 7:
            bucket['stale'] += 1

    return districts


def _build_login_activity_by_district(days=30):
    """
    Har bir District bo'yicha, unga biriktirilgan obyektlarning attached_person
    (Person.profile -> User) xodimlari kesimida so'nggi `days` kunlik tizimga
    kirishlar sonini kunlik kesimda hisoblaydi.

    Qaytadi:
        date_list — kunlar ro'yxati
        district_data — OrderedDict:
            district_key -> {
                'name': str,
                'employees': [
                    {'name': str, 'daily': [int, ...]}, ...
                ],
                'daily_total': [int, ...],   # tuman bo'yicha jami
            }
    """
    today = timezone.now().date()
    start_date = today - datetime.timedelta(days=days - 1)

    # district_key -> {user_id: display_name}
    district_employees = defaultdict(dict)
    district_names = {}

    objects = ConstructionObject.objects.select_related(
        'neighborhood__district', 'attached_person__profile'
    ).filter(
        attached_person__isnull=False,
        attached_person__profile__isnull=False,
    )

    for obj in objects:
        neighborhood = obj.neighborhood
        district = neighborhood.district if neighborhood else None
        district_key = district.id if district else 0
        district_names[district_key] = district.name if district else "Tuman biriktirilmagan"

        person = obj.attached_person
        display_name = person.fullname or person.profile.get_full_name() or person.profile.username
        district_employees[district_key][person.profile_id] = display_name

    all_user_ids = set()
    for employees in district_employees.values():
        all_user_ids |= set(employees.keys())

    login_counts = (
        LoginAttempt.objects.filter(
            user_id__in=all_user_ids,
            timestamp__date__gte=start_date,
            timestamp__date__lte=today,
        )
        .values('user_id', 'timestamp__date')
        .annotate(count=Count('id'))
    )

    login_map = defaultdict(int)
    for row in login_counts:
        login_map[(row['user_id'], row['timestamp__date'])] += row['count']

    date_list = [start_date + datetime.timedelta(days=i) for i in range(days)]

    district_data = OrderedDict()
    for district_key, name in sorted(district_names.items(), key=lambda item: item[1]):
        employees = []
        daily_total = [0] * days

        for user_id, display_name in sorted(district_employees[district_key].items(), key=lambda item: item[1]):
            daily_counts = [login_map.get((user_id, d), 0) for d in date_list]
            employees.append({'name': display_name, 'daily': daily_counts})
            daily_total = [a + b for a, b in zip(daily_total, daily_counts)]

        district_data[district_key] = {
            'name': name,
            'employees': employees,
            'daily_total': daily_total,
        }

    return date_list, district_data


def generate_construction_summary_excel():
    """Ikki varaqli Excel svod hisobotni yaratadi va HttpResponse qaytaradi."""
    wb = Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    subtotal_font = Font(bold=True)
    subtotal_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    warn_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---------------- Sheet 1: Obyektlar svodi ----------------
    ws1 = wb.active
    ws1.title = "Obyektlar svodi"

    headers1 = [
        "Tuman", "Jami obyektlar soni",
        "Moliyalashtirish past (<15%)", "7 kundan ortiq ma'lumot kiritilmagan",
        "Umuman ma'lumot kiritilmagan",
    ]
    ws1.append(headers1)
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    districts = _build_objects_summary_by_district()

    row_idx = 2
    grand_total = grand_low = grand_stale = grand_no_data = 0

    for district_key, district in districts.items():
        ws1.append([
            district['name'], district['total_buildings'],
            district['low_financing'], district['stale'], district['no_data'],
        ])
        for col in range(1, len(headers1) + 1):
            ws1.cell(row=row_idx, column=col).border = border
        if district['low_financing'] or district['stale'] or district['no_data']:
            for col in (3, 4, 5):
                ws1.cell(row=row_idx, column=col).fill = warn_fill
        row_idx += 1

        grand_total += district['total_buildings']
        grand_low += district['low_financing']
        grand_stale += district['stale']
        grand_no_data += district['no_data']

    ws1.append(["JAMI", grand_total, grand_low, grand_stale, grand_no_data])
    for col in range(1, len(headers1) + 1):
        cell = ws1.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.border = border

    for i, width in enumerate([28, 18, 22, 26, 22], start=1):
        ws1.column_dimensions[get_column_letter(i)].width = width
    ws1.freeze_panes = "A2"

    # ---------------- Sheet 2: Kirishlar statistikasi ----------------
    ws2 = wb.create_sheet("Kirishlar statistikasi")
    date_list, district_data = _build_login_activity_by_district(days=30)

    headers2 = ["Tuman", "Hodim"] + [d.strftime("%d.%m") for d in date_list] + ["Jami"]
    ws2.append(headers2)
    for col in range(1, len(headers2) + 1):
        cell = ws2.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    row_idx = 2
    for district_key, district in district_data.items():
        # tuman bo'yicha jami qator
        total_row = [district['name'], "Tuman bo'yicha jami"] + district['daily_total'] + [sum(district['daily_total'])]
        ws2.append(total_row)
        for col in range(1, len(total_row) + 1):
            cell = ws2.cell(row=row_idx, column=col)
            cell.font = subtotal_font
            cell.fill = subtotal_fill
            cell.border = border
        row_idx += 1

        # har bir hodim bo'yicha qator
        for employee in district['employees']:
            row = ["", employee['name']] + employee['daily'] + [sum(employee['daily'])]
            ws2.append(row)
            for col in range(1, len(row) + 1):
                ws2.cell(row=row_idx, column=col).border = border
            row_idx += 1

    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 28
    for i in range(3, len(headers2) + 1):
        ws2.column_dimensions[get_column_letter(i)].width = 9
    ws2.freeze_panes = "C2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"qurilish_svod_hisobot_{timezone.now().date().isoformat()}.xlsx"
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


class UserResource(resources.ModelResource):
    class Meta:
        model = User
        fields = ('username', 'first_name', 'last_name', 'password')
        export_order = fields

@admin.register(User)
class CustomUserAdmin(BaseUserAdmin, ModelAdmin, ImportExportModelAdmin):
    resource_classes = [UserResource]
    form = UserChangeForm
    add_form = UserCreationForm
    change_password_form = AdminPasswordChangeForm
    change_list_template = 'admin/auth/user/change_list.html'
    list_display = ('username', 'first_name', 'last_name', 'role')
    list_display_links = ('username', 'first_name', 'last_name', 'role')

    def get_export_resource_kwargs(self, request, *args, **kwargs):
        kwargs['encoding'] = 'utf-8'
        print('ENCODING')
        return kwargs

    def get_urls(self):
        """Injects custom generation route into default user admin URLs."""
        urls = super().get_urls()
        custom_urls = [
            path(
                'generate-users/',
                self.admin_site.admin_view(self.generate_users_view),
                name='generate_users'
            ),
        ]
        return custom_urls + urls

    def generate_users_view(self, request):
        if request.method == 'POST':
            form = GenerateUsersForm(request.POST)
            if form.is_valid():
                num_users = form.cleaned_data['num_users']

                # 1. Create an in-memory text buffer for the CSV
                buffer = io.StringIO()
                writer = csv.writer(buffer)

                # Write CSV header row
                writer.writerow(['Username', 'Email', 'First Name', 'Last Name'])

                # 2. Build the users list in memory first (extremely fast)
                users_to_create = []
                csv_rows = []

                for _ in range(num_users):
                    unique_id = uuid.uuid4().hex[:8]
                    username = f"user_{unique_id}"
                    email = f"{username}@example.com"
                    first_name = f"First_{unique_id}"
                    last_name = f"Last_{unique_id}"

                    # Instantiate User objects without saving to database yet
                    user_obj = User(
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=last_name,
                    )
                    # Set standard unusable or dummy encrypted password
                    password = get_random_string(14)
                    user_obj.set_password(password)

                    users_to_create.append(user_obj)
                    csv_rows.append([username, email, first_name, last_name, password])

                # 3. Save ALL users to the database in a SINGLE database query
                User.objects.bulk_create(users_to_create)

                # 4. Write all rows to the CSV buffer
                writer.writerows(csv_rows)

                # 5. Return the file download response
                response = HttpResponse(buffer.getvalue(), content_type="text/csv")
                response['Content-Disposition'] = f'attachment; filename="generated_users_{num_users}.csv"'
                return response
        else:
            form = GenerateUsersForm()

        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'opts': self.model._meta,
        }
        return render(request, "admin/generate_users.html", context)

    def get_fieldsets(self, request, obj=None):
        fieldsets = super().get_fieldsets(request, obj)
        fieldsets[0][1]['fields'] = {*fieldsets[0][1]['fields'], 'role'}
        return fieldsets

admin.site.unregister(Group)


@admin.register(LoginAttempt)
class LoginAttemptAdmin(ModelAdmin, ImportExportModelAdmin):
    list_display = ['user', 'ip_address', 'timestamp', 'successful']
    list_filter = ['successful', 'timestamp', 'user', ]
    search_fields = ['user__username', 'ip_address']
    readonly_fields = ['user', 'ip_address', 'user_agent', 'timestamp', 'successful']
    resource_classes = [LoginAttemptsResource]

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False



@admin.register(Group)
class GroupAdmin(BaseGroupAdmin, ModelAdmin):
    pass


class ReviewReportInline(StackedInline):
    model = Report
    extra = 0
    tab = True


class ReviewIssueInline(StackedInline):
    model = Issue
    extra = 0
    tab = True


class ReviewIssuePhotoInline(StackedInline):
    model = IssuePhoto
    extra = 0
    tab = True
    readonly_fields = ['photo_preview']

    def photo_preview(self, obj):
        return mark_safe('<img src="{url}" width="100px" height="100px" />'.format(
            url=obj.photo.url,
        )
        )


class ReviewReportPhotoInline(StackedInline):
    model = ReportPhoto
    extra = 0


@admin.register(Review)
class ReviewAdmin(ModelAdmin):
    inlines = [ReviewIssueInline, ReviewReportInline]


@admin.register(ConstructionObjectDocumentType)
class ConstructionObjectAdminDocumentType(ModelAdmin):
    pass


class DocumentInline(StackedInline):
    model = ConstructionObjectDocument
    extra = 0


@admin.register(Person)
class PersonAdmin(ModelAdmin):
    search_fields = ['fullname']


@admin.register(IssueType)
class IssueTypeAdmin(ModelAdmin):
    pass


@admin.register(InspectionType)
class InspectionTypeAdmin(ModelAdmin):
    pass


@admin.register(ProjectOwnerCompany)
class ProjectOwnerCompanyAdmin(ModelAdmin):
    search_fields = ["name"]

@admin.register(ConstructionCompany)
class ConstructionCompanyAdmin(ModelAdmin):
    search_fields = ["name"]

@admin.register(ProjectDeveloperCompany)
class ProjectDeveloperCompanyAdmin(ModelAdmin):
    search_fields = ["name"]

@admin.register(ConstructionObject)
class ConstructionObjectAdmin(ModelAdmin):
    search_fields = ['name']
    autocomplete_fields = ['attached_person']
    inlines = [DocumentInline]
    form = ConstructionObjectForm
    actions_list = ['excel_summary_report']
    list_display = (
        'id', 'name', 'created_at', 'updated_at', 'address',
        'neighborhood', 'category', 'building_count',
        'p_reviews_p_m', 'i_reviews_p_m', 't_reviews_p_m'
    )
    list_editable = ('category', 'building_count')
    list_display_links = ('id', 'name',)
    list_filter = (
        ('program', RelatedDropdownFilter),
        ('neighborhood', RelatedDropdownFilter),
        ('category', ChoicesDropdownFilter),
        ('owner_companies', RelatedDropdownFilter),
        ('construction_companies', RelatedDropdownFilter),
        ('project_companies', RelatedDropdownFilter),
        ('attached_person', RelatedDropdownFilter),
    )
    list_filter_submit = True
    filter_vertical = ('owner_companies', 'construction_companies', 'project_companies',)
    ordering = ('id', 'name')

    @action(
        description="Excel svod hisobot",
        url_path="excel-summary-report",
    )
    def excel_summary_report(self, request):
        """
        Tuman/Mahalla kesimida obyektlar svodi va so'nggi 30 kunlik
        kirish statistikasini o'z ichiga olgan Excel faylni yuklab beradi.
        """
        return generate_construction_summary_excel()


@admin.register(Issue)
class IssueAdmin(ModelAdmin):
    inlines = [ReviewIssuePhotoInline]


@admin.register(Report)
class ReportAdmin(ModelAdmin):
    inlines = [ReviewReportPhotoInline]


class IssueActionPhotoInline(StackedInline):
    model = IssueActionPhoto
    extra = 0


@admin.register(IssueAction)
class IssueActionAdmin(ModelAdmin):
    inlines = [IssueActionPhotoInline]


class ReviewCommentPhotoInline(StackedInline):
    model = ReviewCommentPhoto
    extra = 0


@admin.register(ReviewComment)
class ReviewCommentAdmin(ModelAdmin):
    inlines = [ReviewCommentPhotoInline]


@admin.register(GovermentProgram)
class GovernmentProgramAdmin(ModelAdmin):
    pass


@admin.register(Region)
class RegionAdmin(ModelAdmin):
    pass


@admin.register(District)
class DistrictAdmin(ModelAdmin):
    autocomplete_fields = ["personal"]


@admin.register(Neighborhood)
class NeighborhoodAdmin(ModelAdmin):
    list_display = ['id', 'name', 'district']
    list_filter = ('district',)
    list_editable = ('district',)


class PublicIssuePhotosInline(StackedInline):
    model = PublicIssuePhoto
    extra = 0


@admin.register(PublicIssue)
class PublicIssueAdmin(ModelAdmin):
    inlines = [PublicIssuePhotosInline, ]
    list_display = ['id', 'title', 'description', ]


@admin.register(ConstructionFinancing)
class ConstructionFinancingAdmin(ModelAdmin):
    list_display = ['id', 'construction', 'date', 'amount', 'person']
    autocomplete_fields = ['construction']

    def get_changeform_initial_data(self, request):
        """Sets the logged-in auth user as the default form value."""
        initial = super().get_changeform_initial_data(request)

        # 'user_field' should match the name of the ForeignKey/Field in your Model
        initial['person'] = request.user.person

        return initial


@admin.register(ConstructionDailyProgress)
class ConstructionDailyProgressAdmin(ModelAdmin):
    list_display = ['id', 'construction', 'date', 'amount', 'workers', 'machines']
    autocomplete_fields = ['construction']


class AssignmentAttachmentInline(StackedInline):
    model = AssignmentAttachment
    extra = 0

@admin.register(Assignment)
class AssignmentAdmin(ModelAdmin):
    list_display = ('id', 'title', 'description', 'created_by', 'assigned_to', 'deadline', 'status')
    list_display_links = ('id', 'title', 'description', 'created_by', 'assigned_to', 'deadline', 'status')
    inlines = [AssignmentAttachmentInline]